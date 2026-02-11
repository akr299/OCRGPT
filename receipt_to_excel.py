#!/usr/bin/env python3
"""
Receipt OCR desktop app (Tkinter).

Features:
- Select a folder of receipt images
- OCR with pytesseract
- Extract structured JSON via OpenAI
- Validate schema
- Append rows into an existing Excel file
- Live progress + scrollable logs

This file is structured for PyInstaller compatibility with a clear `main()` entry point.
"""

from __future__ import annotations

import json
import os
import queue
import re
import threading
import traceback
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

import pytesseract
import tkinter as tk
from openai import OpenAI
from openpyxl import load_workbook
from tkinter import filedialog, messagebox, scrolledtext

# Expected output schema
REQUIRED_SCHEMA = {
    "store": str,
    "date": str,  # YYYY-MM-DD
    "total": int,
    "tax8": int,
    "tax10": int,
    "payment": str,
    "category": str,
}

SUPPORTED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}


def extract_json_from_text(text: str) -> Dict[str, Any]:
    """Extract a JSON object from model text output."""
    cleaned = text.strip()

    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
    if not match:
        raise ValueError("Model response did not contain a JSON object.")

    return json.loads(match.group(0))


def validate_record(record: Dict[str, Any]) -> Dict[str, Any]:
    """Validate and normalize extracted record."""
    missing = [k for k in REQUIRED_SCHEMA if k not in record]
    if missing:
        raise ValueError(f"Missing keys: {missing}")

    extra = [k for k in record if k not in REQUIRED_SCHEMA]
    if extra:
        raise ValueError(f"Unexpected keys: {extra}")

    normalized: Dict[str, Any] = {}

    for key, expected_type in REQUIRED_SCHEMA.items():
        value = record[key]

        if expected_type is int:
            if isinstance(value, int):
                normalized[key] = value
            elif isinstance(value, str):
                digits = re.sub(r"[^0-9-]", "", value)
                if digits in ("", "-"):
                    raise ValueError(f"Key '{key}' is not a valid integer: {value!r}")
                normalized[key] = int(digits)
            else:
                raise ValueError(f"Key '{key}' expected int, got {type(value).__name__}")
        else:
            if not isinstance(value, str):
                raise ValueError(f"Key '{key}' expected str, got {type(value).__name__}")
            normalized[key] = value.strip()

    normalized["date"] = normalized["date"].replace("/", "-").replace(".", "-")
    try:
        datetime.strptime(normalized["date"], "%Y-%m-%d")
    except ValueError as exc:
        raise ValueError(
            f"Invalid date format: {normalized['date']!r}. Expected YYYY-MM-DD"
        ) from exc

    return normalized


def ocr_image(image_path: Path, lang: str = "jpn+eng") -> str:
    """Run OCR for one image."""
    try:
        return pytesseract.image_to_string(str(image_path), lang=lang)
    except Exception as exc:
        raise RuntimeError(f"OCR failed for '{image_path.name}': {exc}") from exc


def call_openai_extract(client: OpenAI, model: str, ocr_text: str) -> Dict[str, Any]:
    """Call OpenAI and return validated JSON record."""
    system_prompt = (
        "You extract structured receipt data. "
        "Return ONLY one JSON object with exactly these keys: "
        "store (string), date (YYYY-MM-DD string), total (integer), tax8 (integer), "
        "tax10 (integer), payment (string), category (string). "
        "Do not include markdown or explanations."
    )

    user_prompt = (
        "Extract data from OCR text below.\n"
        "Rules:\n"
        "- date must be YYYY-MM-DD\n"
        "- total/tax8/tax10 are integers (yen)\n"
        "- if tax8 or tax10 missing, set 0\n"
        "- payment examples: cash, credit_card, ic_card, quickpay, mobile_pay\n"
        "- category examples: food, transport, office, shopping\n\n"
        f"OCR text:\n{ocr_text}"
    )

    try:
        response = client.chat.completions.create(
            model=model,
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
    except Exception as exc:
        raise RuntimeError(f"OpenAI API request failed: {exc}") from exc

    content = response.choices[0].message.content or ""
    parsed = extract_json_from_text(content)
    return validate_record(parsed)


def process_receipt_image(
    image_path: str | Path,
    *,
    client: OpenAI,
    model: str = "gpt-4.1-mini",
    ocr_lang: str = "jpn+eng",
) -> Dict[str, Any]:
    """
    Process one receipt image and return normalized structured record.

    Required by spec:
      process_receipt_image(image_path) -> dict
    Additional dependencies are provided via keyword-only args.
    """
    image_path = Path(image_path)
    ocr_text = ocr_image(image_path, lang=ocr_lang)
    return call_openai_extract(client=client, model=model, ocr_text=ocr_text)


def append_to_excel(record: Dict[str, Any], excel_path: str | Path, sheet_name: str | None = None) -> None:
    """
    Append one validated record to existing Excel file.

    Required by spec:
      append_to_excel(record, excel_path)
    """
    excel_path = Path(excel_path)
    wb = load_workbook(excel_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    keys = list(REQUIRED_SCHEMA.keys())
    header_row = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    header_map = {str(v).strip(): idx + 1 for idx, v in enumerate(header_row) if v is not None}
    has_all_headers = all(k in header_map for k in keys)

    if has_all_headers:
        next_row = ws.max_row + 1
        for key in keys:
            ws.cell(row=next_row, column=header_map[key], value=record[key])
    else:
        ws.append([record[key] for key in keys])

    wb.save(excel_path)


def collect_image_files(input_folder: str | Path) -> List[Path]:
    """Collect image files from a folder (non-recursive)."""
    input_folder = Path(input_folder)
    return sorted(
        p for p in input_folder.iterdir() if p.is_file() and p.suffix.lower() in SUPPORTED_EXTENSIONS
    )


class ReceiptApp(tk.Tk):
    """Tkinter desktop UI for the receipt processing flow."""

    def __init__(self) -> None:
        super().__init__()
        self.title("Receipt OCR to Excel")
        self.geometry("860x620")

        self.folder_var = tk.StringVar()
        self.excel_var = tk.StringVar(value="expense.xlsx")
        self.model_var = tk.StringVar(value="gpt-4.1-mini")
        self.ocr_lang_var = tk.StringVar(value="jpn+eng")
        self.sheet_var = tk.StringVar(value="")
        self.tesseract_var = tk.StringVar(value="")
        self.progress_var = tk.StringVar(value="Progress: 0/0")

        self._queue: queue.Queue = queue.Queue()
        self._worker: threading.Thread | None = None
        self._build_ui()
        self.after(100, self._poll_queue)

    def _build_ui(self) -> None:
        pad = {"padx": 8, "pady": 6}

        tk.Label(self, text="Receipt Image Folder").grid(row=0, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.folder_var, width=70).grid(row=0, column=1, sticky="we", **pad)
        tk.Button(self, text="Browse...", command=self._pick_folder).grid(row=0, column=2, **pad)

        tk.Label(self, text="Excel File (existing)").grid(row=1, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.excel_var, width=70).grid(row=1, column=1, sticky="we", **pad)
        tk.Button(self, text="Browse...", command=self._pick_excel).grid(row=1, column=2, **pad)

        tk.Label(self, text="Model").grid(row=2, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.model_var, width=30).grid(row=2, column=1, sticky="w", **pad)

        tk.Label(self, text="OCR Lang").grid(row=3, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.ocr_lang_var, width=30).grid(row=3, column=1, sticky="w", **pad)

        tk.Label(self, text="Sheet Name (optional)").grid(row=4, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.sheet_var, width=30).grid(row=4, column=1, sticky="w", **pad)

        tk.Label(self, text="Tesseract Path (optional)").grid(row=5, column=0, sticky="w", **pad)
        tk.Entry(self, textvariable=self.tesseract_var, width=70).grid(row=5, column=1, sticky="we", **pad)

        self.run_button = tk.Button(self, text="Run", command=self._on_run_clicked, width=16)
        self.run_button.grid(row=6, column=1, sticky="w", **pad)

        tk.Label(self, textvariable=self.progress_var, fg="blue").grid(row=6, column=1, sticky="e", **pad)

        tk.Label(self, text="Log").grid(row=7, column=0, sticky="nw", **pad)
        self.log_text = scrolledtext.ScrolledText(self, width=100, height=24, state="disabled")
        self.log_text.grid(row=7, column=1, columnspan=2, sticky="nsew", **pad)

        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(7, weight=1)

    def _pick_folder(self) -> None:
        folder = filedialog.askdirectory(title="Select receipt image folder")
        if folder:
            self.folder_var.set(folder)

    def _pick_excel(self) -> None:
        file_path = filedialog.askopenfilename(
            title="Select existing Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.xltx *.xltm"), ("All files", "*.*")],
        )
        if file_path:
            self.excel_var.set(file_path)

    def _log(self, text: str) -> None:
        self.log_text.configure(state="normal")
        self.log_text.insert(tk.END, text + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state="disabled")

    def _on_run_clicked(self) -> None:
        if self._worker and self._worker.is_alive():
            return

        if not os.getenv("OPENAI_API_KEY"):
            messagebox.showerror("Missing API Key", "OPENAI_API_KEY is not set. Please set it before running.")
            return

        folder = Path(self.folder_var.get().strip())
        excel = Path(self.excel_var.get().strip())

        if not folder.exists() or not folder.is_dir():
            messagebox.showerror("Invalid Folder", "Please select a valid receipt image folder.")
            return

        if not excel.exists():
            messagebox.showerror("Invalid Excel File", "Please select an existing Excel file.")
            return

        images = collect_image_files(folder)
        if not images:
            messagebox.showerror("No Images Found", "No supported image files were found in the selected folder.")
            return

        self.run_button.configure(state="disabled")
        self.progress_var.set(f"Progress: 0/{len(images)}")
        self._log("=== Started ===")
        self._log(f"Folder: {folder}")
        self._log(f"Excel : {excel}")

        self._worker = threading.Thread(
            target=self._worker_run,
            kwargs={
                "images": images,
                "excel_path": excel,
                "model": self.model_var.get().strip() or "gpt-4.1-mini",
                "ocr_lang": self.ocr_lang_var.get().strip() or "jpn+eng",
                "sheet_name": self.sheet_var.get().strip() or None,
                "tesseract_cmd": self.tesseract_var.get().strip() or None,
            },
            daemon=True,
        )
        self._worker.start()

    def _worker_run(
        self,
        *,
        images: List[Path],
        excel_path: Path,
        model: str,
        ocr_lang: str,
        sheet_name: str | None,
        tesseract_cmd: str | None,
    ) -> None:
        if tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

        client = OpenAI()
        total = len(images)
        success = 0
        failures: List[str] = []

        for idx, image_path in enumerate(images, start=1):
            try:
                record = process_receipt_image(
                    image_path,
                    client=client,
                    model=model,
                    ocr_lang=ocr_lang,
                )
                append_to_excel(record, excel_path, sheet_name=sheet_name)
                success += 1
                self._queue.put(("log", f"[OK] {image_path.name} -> {record}"))
            except Exception as exc:
                msg = f"{image_path.name}: {exc}"
                failures.append(msg)
                self._queue.put(("log", f"[NG] {msg}"))

            self._queue.put(("progress", idx, total))

        self._queue.put(("done", success, total, failures))

    def _poll_queue(self) -> None:
        try:
            while True:
                item = self._queue.get_nowait()
                kind = item[0]

                if kind == "log":
                    self._log(item[1])
                elif kind == "progress":
                    current, total = item[1], item[2]
                    self.progress_var.set(f"Progress: {current}/{total}")
                elif kind == "done":
                    success, total, failures = item[1], item[2], item[3]
                    self.run_button.configure(state="normal")
                    self._log("=== Finished ===")
                    self._log(f"Succeeded: {success}/{total}")

                    if failures:
                        self._log("Failures:")
                        for f in failures:
                            self._log(f"- {f}")
                        messagebox.showerror(
                            "OCR / API Failure",
                            "Some receipts failed during OCR or API extraction.\n"
                            f"Succeeded: {success}/{total}\n"
                            f"Failed: {len(failures)}",
                        )
                    else:
                        messagebox.showinfo("Completed", f"All receipts processed successfully ({success}/{total}).")
        except queue.Empty:
            pass
        except Exception:
            self.run_button.configure(state="normal")
            messagebox.showerror("Unexpected Error", traceback.format_exc())

        self.after(100, self._poll_queue)


def main() -> None:
    """Main entry point for desktop app + PyInstaller."""
    app = ReceiptApp()
    app.mainloop()


if __name__ == "__main__":
    main()
