from __future__ import annotations

import json
import re
from dataclasses import dataclass, field
from datetime import date
from pathlib import Path
from typing import Any, Dict, List, Sequence, Tuple
from urllib.parse import quote

import pytesseract
from openai import AuthenticationError, OpenAI, RateLimitError
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

SUPPORTED_EXTENSIONS = {".png", ".jpg", ".jpeg", ".bmp", ".tif", ".tiff", ".webp"}

EXCEL_COLUMNS = [
    "日付",
    "勘定科目",
    "内容",
    "支払先",
    "支払方法",
    "金額(税込)",
    "税区分",
    "事業利用割合(%)",
    "経費計上額",
    "備考",
    "is_error",
    "error_reason",
]

FIELD_ALIASES = {
    "date": "日付",
    "category": "勘定科目",
    "description": "内容",
    "content": "内容",
    "store": "支払先",
    "vendor": "支払先",
    "payment": "支払方法",
    "total": "金額(税込)",
    "tax": "税区分",
    "business_ratio": "事業利用割合(%)",
    "expense_amount": "経費計上額",
    "note": "備考",
}

JAPANESE_ACCOUNTING_CATEGORIES = [
    "旅費交通費",
    "会議費",
    "接待交際費",
    "消耗品費",
    "通信費",
    "水道光熱費",
    "広告宣伝費",
    "新聞図書費",
    "地代家賃",
    "外注工賃",
    "支払手数料",
    "租税公課",
    "雑費",
]

ERA_YEAR_OFFSET = {
    "令和": 2018,
    "平成": 1988,
    "昭和": 1925,
    "大正": 1911,
    "明治": 1867,
}

ERROR_ROW_FILL = PatternFill(fill_type="solid", start_color="FFF4B4B4", end_color="FFF4B4B4")


@dataclass
class ReceiptRecord:
    file_name: str
    source_image_path: str
    fields: Dict[str, Any] = field(default_factory=dict)
    is_error: bool = False
    error_reason: str = ""

    def get(self, key: str, default: Any = "") -> Any:
        return self.fields.get(key, default)

    def set(self, key: str, value: Any) -> None:
        self.fields[key] = value

    def mark_error(self, reason: str) -> None:
        self.is_error = True
        existing = [item for item in self.error_reason.split("|") if item]
        if reason and reason not in existing:
            existing.append(reason)
        self.error_reason = "|".join(existing)

    def clear_error(self) -> None:
        self.is_error = False
        self.error_reason = ""

    def source_image_link(self) -> str:
        resolved = Path(self.source_image_path).resolve().as_posix()
        return f"file:///{quote(resolved)}"

    def excel_row(self, columns: Sequence[str]) -> List[Any]:
        row_values: List[Any] = []
        for col in columns:
            if col == "is_error":
                row_values.append(1 if self.is_error else 0)
            elif col == "error_reason":
                row_values.append(self.error_reason)
            elif col == "source_image_link":
                row_values.append(self.source_image_link())
            elif col == "file_name":
                row_values.append(self.file_name)
            else:
                row_values.append(self.get(col, ""))
        return row_values


@dataclass
class ProcessFailure:
    file_name: str
    reason: str


def normalize_date_value(value: Any) -> Tuple[str, str]:
    """
    日付文字列を YYYY/MM/DD に正規化する。
    - 西暦/和暦の変換に成功した場合のみ正規化値を返す
    - 変換不能な場合は元文字列を保持しつつ date_parse_failed を返す
    """
    text = str(value or "").strip()
    if not text:
        return "", ""

    normalized_input = text.replace(".", "/").replace("-", "/")

    seireki_match = re.search(r"^(\d{4})\s*/\s*(\d{1,2})\s*/\s*(\d{1,2})$", normalized_input)
    if seireki_match:
        y, m, d = map(int, seireki_match.groups())
        try:
            return date(y, m, d).strftime("%Y/%m/%d"), ""
        except ValueError:
            return text, "date_parse_failed"

    wareki_match = re.search(
        r"(令和|平成|昭和|大正|明治)\s*(元|\d{1,2})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日",
        text,
    )
    if wareki_match:
        era_name, era_year, month, day = wareki_match.groups()
        era_numeric_year = 1 if era_year == "元" else int(era_year)
        year = ERA_YEAR_OFFSET[era_name] + era_numeric_year
        try:
            return date(year, int(month), int(day)).strftime("%Y/%m/%d"), ""
        except ValueError:
            return text, "date_parse_failed"

    compact_match = re.search(r"^(\d{8})$", re.sub(r"\D", "", text))
    if compact_match:
        digits = compact_match.group(1)
        y, m, d = int(digits[:4]), int(digits[4:6]), int(digits[6:8])
        try:
            return date(y, m, d).strftime("%Y/%m/%d"), ""
        except ValueError:
            return text, "date_parse_failed"

    return text, "date_parse_failed"


def calculate_expense_amount(amount_text: Any, ratio_text: Any) -> str:
    amount_digits = re.sub(r"[^0-9-]", "", str(amount_text or ""))
    ratio_raw = str(ratio_text or "").strip().replace("％", "%").replace("%", "")
    if not amount_digits or amount_digits == "-" or not ratio_raw:
        return ""

    try:
        amount = float(amount_digits)
        ratio = float(ratio_raw)
    except ValueError:
        return ""

    return str(int(round(amount * ratio / 100.0)))


def infer_tax_category(value: Any) -> Tuple[str, str]:
    """
    税区分は OCR/AI 文字列から税率が判別できる場合のみ返す。
    8% / 10% を優先し、不明時は空文字 + tax_rate_unknown を返す。
    """
    text = str(value or "").strip()
    if not text:
        return "", "tax_rate_unknown"

    normalized = text.replace("％", "%").replace("１０", "10").replace("８", "8")
    if re.search(r"\b10\s*%", normalized) or "10%" in normalized:
        return "10%", ""
    if re.search(r"\b8\s*%", normalized) or "8%" in normalized:
        return "8%", ""
    if "軽減" in normalized:
        return "8%", ""

    return "", "tax_rate_unknown"


def _extract_json_from_text(text: str) -> Dict[str, Any]:
    cleaned = text.strip()
    if cleaned.startswith("```"):
        cleaned = re.sub(r"^```(?:json)?\s*", "", cleaned, flags=re.IGNORECASE)
        cleaned = re.sub(r"\s*```$", "", cleaned)

    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        pass

    match = re.search(r"\{.*\}", cleaned, flags=re.DOTALL)
    if not match:
        raise ValueError("OpenAI response did not contain a JSON object")
    return json.loads(match.group(0))


def _normalize_record(raw: Dict[str, Any], image_path: Path) -> ReceiptRecord:
    record = ReceiptRecord(
        file_name=image_path.name,
        source_image_path=str(image_path),
    )

    for key, value in raw.items():
        normalized_key = FIELD_ALIASES.get(key, key)
        if isinstance(value, (dict, list)):
            record.set(normalized_key, json.dumps(value, ensure_ascii=False))
        else:
            record.set(normalized_key, "" if value is None else str(value).strip())

    if not record.get("支払先"):
        record.mark_error("store_not_found")

    total_raw = record.get("金額(税込)", "")
    digits = re.sub(r"[^0-9-]", "", str(total_raw))
    if digits in {"", "-"}:
        record.mark_error("amount_not_found")
        record.set("金額(税込)", "")
    else:
        record.set("金額(税込)", digits)

    normalized_date, date_error = normalize_date_value(record.get("日付", ""))
    record.set("日付", normalized_date)
    if date_error:
        record.mark_error(date_error)

    tax_value, tax_error = infer_tax_category(record.get("税区分", ""))
    record.set("税区分", tax_value)
    if tax_error:
        record.mark_error(tax_error)

    if not record.get("勘定科目"):
        record.set("勘定科目", "雑費")
    if not record.get("支払方法"):
        record.set("支払方法", "不明")

    # 仕様: 以下3項目は OCR/AI 抽出では必ず空白初期化し、GUI手動入力を前提とする。
    record.set("事業利用割合(%)", "")
    record.set("経費計上額", "")
    record.set("備考", "")

    for column in EXCEL_COLUMNS:
        if column in {"is_error", "error_reason"}:
            continue
        record.set(column, str(record.get(column, "")).strip())

    return record


def collect_image_files(input_folder: Path) -> List[Path]:
    return sorted(
        path
        for path in input_folder.iterdir()
        if path.is_file() and path.suffix.lower() in SUPPORTED_EXTENSIONS
    )


class ReceiptProcessor:
    def __init__(
        self,
        api_key: str,
        model: str = "gpt-4.1-mini",
        ocr_lang: str = "jpn+eng",
        tesseract_cmd: str | None = None,
    ):
        self.api_key = api_key
        self.model = model
        self.ocr_lang = ocr_lang
        if tesseract_cmd:
            pytesseract.pytesseract.tesseract_cmd = tesseract_cmd

    def _ocr_image(self, image_path: Path) -> str:
        text = pytesseract.image_to_string(str(image_path), lang=self.ocr_lang)
        if not text.strip():
            raise ValueError("ocr_empty")
        return text

    def _call_openai(self, client: OpenAI, ocr_text: str) -> Dict[str, Any]:
        system_prompt = (
            "You extract structured data from Japanese receipts. "
            "Return ONLY one JSON object. Include keys: 日付, 勘定科目, 内容, 支払先, 支払方法, 金額(税込), 税区分."
        )
        user_prompt = (
            "OCR text から経費精算向けデータを抽出してください。\n"
            "Rules:\n"
            "- 日付 は可能なら日付文字列。和暦でも可\n"
            "- 勘定科目 は日本語の会計カテゴリ\n"
            "- 金額(税込) は金額\n"
            "- 支払方法 は支払方法\n"
            "- 税区分 は税率が分かる場合のみ記載(8%/10%)\n"
            "- OCR原文の全文をフィールドへ出力しない\n"
            "- わからない項目は空文字で返す\n"
            f"OCR:\n{ocr_text}"
        )

        response = client.chat.completions.create(
            model=self.model,
            temperature=0,
            response_format={"type": "json_object"},
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": user_prompt},
            ],
        )
        content = response.choices[0].message.content or ""
        return _extract_json_from_text(content)

    def process_folder(self, input_folder: Path) -> Tuple[List[ReceiptRecord], List[ProcessFailure]]:
        if not input_folder.exists() or not input_folder.is_dir():
            raise FileNotFoundError(f"画像フォルダが見つかりません: {input_folder}")

        image_files = collect_image_files(input_folder)
        if not image_files:
            raise FileNotFoundError("対応画像ファイルがありません。")

        if not self.api_key.strip():
            raise RuntimeError("OpenAI APIキーが設定されていません。設定から再登録してください。")

        client = OpenAI(api_key=self.api_key)

        records: List[ReceiptRecord] = []
        failures: List[ProcessFailure] = []

        for image_path in image_files:
            try:
                ocr_text = self._ocr_image(image_path)
                raw_record = self._call_openai(client, ocr_text)
                record = _normalize_record(raw_record, image_path=image_path)
                records.append(record)
            except AuthenticationError:
                raise RuntimeError("OpenAI APIキー認証に失敗しました。設定からAPIキーを再登録してください。")
            except RateLimitError:
                raise RuntimeError("OpenAIの利用上限に達しました。課金状況を確認し、必要ならAPIキーを再設定してください。")
            except Exception as exc:
                failures.append(ProcessFailure(file_name=image_path.name, reason=str(exc)))
                error_record = ReceiptRecord(
                    file_name=image_path.name,
                    source_image_path=str(image_path),
                    fields={col: "" for col in EXCEL_COLUMNS if col not in {"is_error", "error_reason"}},
                    is_error=True,
                    error_reason=str(exc),
                )
                records.append(error_record)

        return records, failures


def _build_column_order(records: Sequence[ReceiptRecord]) -> List[str]:
    _ = records
    return list(EXCEL_COLUMNS)


def append_records_to_excel(excel_path: Path, records: Sequence[ReceiptRecord], sheet_name: str | None = None) -> int:
    if not records:
        raise ValueError("保存対象データがありません")

    if excel_path.exists():
        wb = load_workbook(excel_path)
        ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
    else:
        wb = Workbook()
        ws = wb.active
        if sheet_name:
            ws.title = sheet_name

    columns = _build_column_order(records)
    ws.delete_rows(1, ws.max_row)
    ws.append(columns)

    for row_index, record in enumerate(records, start=2):
        ws.append(record.excel_row(columns))
        if record.is_error:
            for cell in ws[row_index]:
                cell.fill = ERROR_ROW_FILL

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    wb.save(excel_path)
    return len(records)
